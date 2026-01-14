from PyQt5.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QLabel, QTextEdit,
    QPushButton, QFileDialog, QMessageBox
)
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QFont
import numpy as np
from matplotlib.figure import Figure
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from copy import deepcopy
from scipy.optimize import curve_fit
from scipy.interpolate import interp1d
from datetime import datetime
import zipfile
import io
import os
from openpyxl import load_workbook
import gwyfile
from gwyfile.objects import GwyContainer, GwySIUnit
import json


class calibrationset:
    def __init__(self, data_path, version='v0.5'):
        now = datetime.now()
        self.ident = now.strftime("%d/%m/%Y %H:%M") + "F" + data_path
        self.version = version

    def fill_set_v05(self, cal_name, quality, X_cal, Y_cal, initialguess, res, cc, meas):
        self.sample = cal_name
        self.Dopant = self.Dopant_type
        self.carrier = self.carrier_type
        self.setting = self.cal_setting
        self.dat = np.c_[self.ref(X_cal), Y_cal]
        self.quality = quality
        self.initialguess = initialguess
        self.res = res
        self.cc = cc
        self.meas = meas

    def fill_set(self, *args):
        if self.version == "v0.5":
            self.fill_set_v05(*args)

    def save_to_database(self):
        if self.version == "v0.5":
            package = np.array(
                [self.sample, self.Dopant, self.carrier, self.setting, self.dat,
                 self.quality, self.initialguess, self.res, self.cc, self.meas, self.version],
                dtype=object
            )
            try:
                database_path = r"Z:\2_Reference\calibration_database.npz"
                lead_num = 0
                if os.path.exists(database_path):
                    try:
                        with np.load(database_path, allow_pickle=True) as loaded_npz:
                            lead_num = len(loaded_npz.files)
                    except Exception as e:
                        print(f"Warning: Could not read existing database: {e}. Creating new file.")
                else:
                    print("Database file does not exist. Creating new file.")
                bio = io.BytesIO()
                np.save(bio, package)
                with zipfile.ZipFile(database_path, 'a' if os.path.exists(database_path) else 'w') as zipf:
                    zipf.writestr(f"{lead_num}T{self.ident}.npy", data=bio.getbuffer().tobytes())
                print(f"Successfully saved to {database_path} with identifier {lead_num}T{self.ident}")
            except Exception as e:
                print(f"Error saving to database: {e}")

def save_measurement_settings_to_json(main_window):
    """Save parameters with user-chosen name (or default timestamp)."""

    project_dir = os.path.expanduser("~/AppName/saved_projects")
    os.makedirs(project_dir, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    default_name = f"project_{timestamp}"

    file_path, _ = QFileDialog.getSaveFileName(
        main_window,
        "Save Project Parameters",
        os.path.join(project_dir, default_name),
        "JSON Files (*.json);;All Files (*)"
    )
    if not file_path:
        return  # cancelled

    # Ensure .json extension
    if not file_path.lower().endswith(".json"):
        file_path += ".json"

    meas = main_window.import_measurement_tab
    calib = main_window.select_calibration_tab
    align = main_window.alignment_tab
    fit = main_window.fitpoints_tab

    settings = {
        "project_saved_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "import_measurement": {
            "data_type": meas.ui.dataTypeComboBox.currentText(),
            "measurement_file": getattr(meas, "path_data", "") or "",
            "denomination": meas.ui.denominationLineEdit.text(),
            "flip_data": meas.ui.flipDataCheckBox.isChecked(),
            "left_border_um": float(meas.borders_data[0]),
            "right_border_um": float(meas.borders_data[1]),
        },
        "select_calibration": {
            "Calibration sample": calib.ui.calib_sample_combobox.currentText(),
            "preset": calib.ui.Preset_comboBox.currentText() if not calib.ui.Preset_comboBox.isHidden() else "",
            "linear_scale": calib.ui.RawData_LinearScale_checkBox.isChecked(),
            "data_type": calib.ui.Calib_Data_Type_comboBox.currentText() if calib.ui.Calib_Data_Type_comboBox.isVisible() else "",
            "flip_calibration": calib.ui.Flip_Data_Checkbox.isChecked(),
            "left_border_um": float(calib.borders_data[0]),
            "right_border_um": float(calib.borders_data[1]),
            "denomination": calib.ui.Calib_Data_Denom_lineEdit.text(),
            "dopant_type": calib.ui.Dopant_Type_comboBox.currentText(),
            "number_of_steps": calib.ui.Nb_steps_spinBox.value(),
            "min_step_distance": float(calib.ui.Min_Step_LineEdit.text() or 0),
        },
        "alignment": {
            "filter_strength": align.ui.DataFilterStrenght_slider.value(),
            "filter_order": align.ui.FilterOrder_Slider.value(),
            "min_stretch": align.ui.minStretch_slider.value(),
            "max_stretch": align.ui.MaxStretch_slider.value(),
            "increase_search_area": align.ui.Increase_Search_area_checkbox.isChecked(),
            "shift_resolution": align.ui.Search_resol_shift_lineedit.text(),
            "stretch_resolution": align.ui.Search_resol_Stretch_lineedit.text(),
            "fine-alignment_number_of_evaluated_points": align.ui.fine_alignement_lineedit.text(),
        },
        "fitpoints": {
            "mode": fit.G_fit_Mode,
            "number_of_points": fit.ui.Nbr_of_points_spinBox.value(),
            "min_distance": float(fit.ui.Min_distance_between_steps_lineEdit.text() or 0),
            "include_left_edge": fit.ui.include_left_edge_as_anchor_checkBox.isChecked(),
            "include_right_edge": fit.ui.include_right_edge_as_anchor_checkBox.isChecked(),
            "intermediate_points": [s.value() for s in fit.sliders],
        }
    }

    try:
        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(settings, f, indent=4)
        print(f"Saved: {file_path}")
        QMessageBox.information(main_window, "Success", f"Saved:\n{os.path.basename(file_path)}")
    except Exception as e:
        QMessageBox.critical(main_window, "Error", f"Save failed:\n{e}")
        
class CalibrationTab:
    """Controller for the 'Calibration' tab functionality."""

    def __init__(self, ui, main_window):
        """Initialize the tab with UI elements and data attributes."""
        self.ui = ui
        self.main_window = main_window
        self.alignment_tab = main_window.alignment_tab
        self.import_measurement_tab = main_window.import_measurement_tab
        self.select_calibration_tab = main_window.select_calibration_tab
        self.fitpoints_tab = main_window.fitpoints_tab

        # ---------------------------------------------------------------------
        # Global settings
        # ---------------------------------------------------------------------
        self.G_alignment_fine_iterations = 50
        self.G_electron_const = 1.6E-19
        self.G_max_N = [1E14, 1E22]
        self.export_excel_metadata = False
        # self.XLS = r"C:\Users\allani\Desktop\Internship\Tasks\Task3\LoadNPZ\Template.xlsx"
        self.XLS = r"Z:\2_Reference\Quantification_SAMPLE_PROBE__ID.xlsx"
        self.G_canvas_aspect_ratio = np.array([544, 300])
        self.G_canvas_dpi = 80
        self.calibration_convert_metadata = [False, False]

        # ---------------------------------------------------------------------
        # Button colors (initial state)
        # ---------------------------------------------------------------------
        self.ui.calibration_startt_pushButton.setStyleSheet("background-color: red; color: black")
        self.ui.Convert_to_Charge_Carriers_Button.setStyleSheet("background-color: red; color: black")
        self.ui.Save_To_Database_Button.setStyleSheet("background-color: red; color: black")
        self.ui.Create_excel_File_Button.setStyleSheet("background-color: red; color: black")
        self.ui.Save_As_Png_Button.setStyleSheet("background-color: red; color: black")
        self.ui.Apply_To_Gwyddion_File_Button.setStyleSheet("background-color: red; color: black")

        # ---------------------------------------------------------------------
        # Matplotlib canvases
        # ---------------------------------------------------------------------
        self.figure_calibration_overlay = Figure(
            figsize=1.5 * self.G_canvas_aspect_ratio / self.G_canvas_dpi, dpi=self.G_canvas_dpi
        )
        self.canvas_calibration_overlay = FigureCanvas(self.figure_calibration_overlay)
        try:
            self.ui.calibration_overlay_verticalLayout.addWidget(self.canvas_calibration_overlay)
            print("Calibration overlay canvas added to calibration_overlay_verticalLayout")
        except AttributeError as e:
            print(f"Error: {e}. Ensure calibration_overlay_verticalLayout exists.")
            raise SystemExit(1)

        self.figure_calibration_curve = Figure(
            figsize=1.5 * self.G_canvas_aspect_ratio / self.G_canvas_dpi, dpi=self.G_canvas_dpi
        )
        self.canvas_calibration_curve = FigureCanvas(self.figure_calibration_curve)
        try:
            self.ui.Calibration_Curve_verticalLayout.addWidget(self.canvas_calibration_curve)
            print("Calibration curve canvas added to Calibration_Curve_verticalLayout")
        except AttributeError as e:
            print(f"Error: {e}. Ensure Calibration_Curve_verticalLayout exists.")
            raise SystemExit(1)

        # ---------------------------------------------------------------------
        # Button wiring
        # ---------------------------------------------------------------------
        try:
            self.ui.Convert_to_Charge_Carriers_Button.clicked.connect(self.calibration_convert)
            print("Convert_to_Charge_Carriers_Button signal connected")
        except AttributeError as e:
            print(f"Error: {e}. Ensure Convert_to_Charge_Carriers_Button exists.")
            raise SystemExit(1)

        try:
            self.ui.Save_To_Database_Button.clicked.connect(self.export_database)
            print("Save_To_Database_Button signal connected")
        except AttributeError as e:
            print(f"Error: {e}. Ensure Save_To_Database_Button exists.")
            raise SystemExit(1)

        try:
            self.ui.Create_excel_File_Button.clicked.connect(self.export_excel)
            print("Create_excel_File_Button connected")
        except AttributeError as e:
            print(f"Error: {e}. Ensure Create_excel_File_Button exists in UI.")
            raise SystemExit(1)

        try:
            self.ui.Save_As_Png_Button.clicked.connect(self.save_as_png)
            print("Save_As_Png_Button connected")
        except AttributeError as e:
            print(f"Error: {e}. Ensure Save_As_Png_Button exists in UI.")
            raise SystemExit(1)

        try:
            self.ui.Apply_To_Gwyddion_File_Button.clicked.connect(self.apply_to_gwyddion_file)
            print("Apply_To_Gwyddion_File_Button connected")
        except AttributeError as e:
            print(f"Error: {e}. Ensure Apply_To_Gwyddion_File_Button exists in UI.")
            raise SystemExit(1)

    # -------------------------------------------------------------------------
    # UI helpers
    # -------------------------------------------------------------------------

    def reset_sstart_calib_button_state(self):
        """Reset measurement import state and update button colors."""
        self.ui.calibration_startt_pushButton.setStyleSheet("background-color: yellow; color: black")

    def reset_buttons_state(self):
        """Update button colors."""
        self.ui.calibration_startt_pushButton.setStyleSheet("background-color: red; color: black")
        self.ui.Convert_to_Charge_Carriers_Button.setStyleSheet("background-color: red; color: black")
        self.ui.Save_To_Database_Button.setStyleSheet("background-color: red; color: black")
        self.ui.Create_excel_File_Button.setStyleSheet("background-color: red; color: black")
        self.ui.Save_As_Png_Button.setStyleSheet("background-color: red; color: black")
        self.ui.Apply_To_Gwyddion_File_Button.setStyleSheet("background-color: red; color: black")

    def draw_xlabel(self, Quantity="Depth", is_log=False, figure=None):
        """Set x-axis label for the specified figure."""
        figure = figure or self.figure_calibration_curve
        ax = figure.gca() if figure.axes else figure.add_subplot(111)
        if Quantity == "Depth":
            ax.set_xlabel("Depth [mm]")
        else:
            label = (
                'SSRM measured resistance'
                if self.main_window.import_measurement_tab.G_dat_datatype == "SSRM"
                else self.main_window.import_measurement_tab.G_dat_denomination
            )
            if is_log and self.main_window.import_measurement_tab.G_dat_datatype == "SSRM":
                label += ' [$log_{10}(\Omega)$]'
            elif self.main_window.import_measurement_tab.G_dat_datatype == "SSRM":
                label += ' [$\Omega$]'
            ax.set_xlabel(label, fontsize=10)
            ax.tick_params(axis='both', which='major', labelsize=10)
            ax.tick_params(axis='both', which='minor', labelsize=10)

    def draw_ylabel(self, Quantity="Calibration", is_log=True, figure=None):
        """Set y-axis label for the specified figure."""
        # kept fallback consistent with your pattern; you always pass figure in calls below
        figure = figure or self.figure_calibration_curve
        ax = figure.gca()
        if Quantity == "Calibration":
            if self.main_window.select_calibration_tab.G_cal_setting == 1:
                carrier_label = "p-type" if self.main_window.select_calibration_tab.G_carrier_type == "B" else "n-type"
                identifier = f"{carrier_label} charge carrier concentration"
                unit = "cm$^{-3}$"
            elif self.main_window.select_calibration_tab.G_cal_setting == 2:
                identifier = "SRP measured resistivity ρ"
                unit = "Ωcm"
            elif self.main_window.select_calibration_tab.G_cal_setting == 3:
                ax.set_ylabel(self.main_window.select_calibration_tab.denomination, fontsize=10)
                print(f"Y label set to {self.main_window.select_calibration_tab.denomination}")
                return
            if is_log:
                label = f"{identifier} [$log_{{10}}$({unit})]"
            else:
                label = f"{identifier} [{unit}]"
        elif Quantity == "Data":
            # This branch mirrors your original; not used by this tab in practice.
            label = "SSRM measured resistance" if getattr(self, "G_dat_datatype", "SSRM") == "SSRM" else self.import_measurement_tab.G_dat_denomination
            if is_log and getattr(self, "G_dat_datatype", "SSRM") == "SSRM":
                label += " [$log_{10}(\Omega)$]"
            elif getattr(self, "G_dat_datatype", "SSRM") == "SSRM":
                label += " [$\Omega$]"
        ax.set_ylabel(label, fontsize=10)
        print(f"Y label set to {label}")

    def draw_grid(self, ax):
        """Add grid to the plot."""
        ax.grid(color='b', which='minor', linestyle='-.', linewidth=0.25)
        ax.grid(color='b', which='major', linestyle='-.', linewidth=0.5)

    # -------------------------------------------------------------------------
    # Plot initializers
    # -------------------------------------------------------------------------

    def redraw_calibration_overlay_init(self, ax, X_cal, Y_cal, X_dat, initialguess_calibrated):
        """Initialize calibration overlay plot."""
        ax.set_title("Wrapping Measurement Profile over Linear Interpolation")
        ax.plot(X_cal, np.power(10., Y_cal), color='k', ls='-', label='Calibration Data')
        ax.plot(
            X_dat, np.power(10., initialguess_calibrated), color='blue', ls='--',
            label='Measurement Data Calibrated with Initial Guess', zorder=0
        )
        ax.set_yscale('log')
        self.draw_grid(ax)
        ax.legend(loc='best', fontsize=10)

    def redraw_calibration_curve_init(self, ax, X_cal, Y_cal, Y_plateaus_dat, Y_plateaus_cal, mode="Calibration"):
        """Initialize calibration curve plot."""
        if mode == "Calibration":
            ax.plot(np.power(10., Y_plateaus_dat), np.power(10., Y_plateaus_cal),
                    color='blue', label='calibration (initial guess)', zorder=0)
        ax.scatter(np.power(10., self.alignment_tab.ref(X_cal)), np.power(10., Y_cal),
                   color='blue', alpha=0.25, label='Datapoints')
        self.draw_grid(ax)
        ax.set_xscale('log')
        ax.set_yscale('log')
        ax.legend(loc='best', fontsize=10)

    # -------------------------------------------------------------------------
    # Physics helpers
    # -------------------------------------------------------------------------

    def mobility_masetti(self, N, Dopant_type):
        """Compute mobility using Masetti model."""
        mu_0 = [52.2, 68.5, 44.9]
        mu_max = [1417, 1414, 470.5]
        mu_1 = [43.4, 56.1, 29]
        C_r = [9.68E16, 9.20E16, 2.23E17]
        C_s = [3.43E20, 3.41E20, 6.1E20]
        alpha = [0.680, 0.711, 0.719]
        beta = [2.00, 1.98, 2.00]
        if Dopant_type == 'As':
            i = 0
        elif Dopant_type == 'P':
            i = 1
        elif Dopant_type == 'B':
            i = 2
        else:
            print('Dopant type must be B, P, or As')
            return None

        if i < 2:
            mu = mu_0[i] + (mu_max[i] - mu_0[i]) / (1 + np.power(N / C_r[i], alpha[i])) - mu_1[i] / (1 + np.power(C_s[i] / N, beta[i]))
        else:
            P_c = 9.23E16
            mu = mu_0[i] * np.exp(-P_c / N) + mu_max[i] / (1 + np.power(N / C_r[i], alpha[i])) - mu_1[i] / (1 + np.power(C_s[i] / N, beta[i]))
        return mu

    def convert_N_to_rho(self, N, Dopant_type):
        """Convert carrier concentration to resistivity."""
        mu = self.mobility_masetti(N, Dopant_type)
        if mu is None:
            return None
        return 1 / (N * mu * self.G_electron_const)

    def convert_rho_to_N(self, array):
        """Convert resistivity to carrier concentration."""
        Dopant_type = self.select_calibration_tab.G_dopant_type
        N_values = np.logspace(np.log10(self.G_max_N[0] * 0.5), np.log10(self.G_max_N[1] * 2), 1000)
        rho_values = self.convert_N_to_rho(N_values, Dopant_type)
        if rho_values is None:
            print(f"Error: convert_N_to_rho returned None for Dopant_type={Dopant_type}")
            return None
        convert_rho_to_N_func = interp1d(
            rho_values, N_values, bounds_error=False,
            fill_value=(self.G_max_N[1] * 2, self.G_max_N[0] * 0.5)
        )
        rho_allowed_interval = self.convert_N_to_rho(np.array(self.G_max_N), Dopant_type)
        if rho_allowed_interval is None:
            print(f"Error: rho_allowed_interval is None for Dopant_type={Dopant_type}")
            return None
        array = np.array(array)
        array = np.clip(array, rho_allowed_interval[1], rho_allowed_interval[0])
        return convert_rho_to_N_func(array)

    # -------------------------------------------------------------------------
    # Main actions
    # -------------------------------------------------------------------------

    def calibration_convert(self):
        """Convert calibration data to charge carrier concentration and update plot."""
        if not (hasattr(self, 'fitpoints_dat_opt') and hasattr(self.fitpoints_tab, 'Y_plateaus_cal') and
                hasattr(self.fitpoints_tab, 'Y_plateaus_dat') and hasattr(self.alignment_tab, 'X_c') and
                hasattr(self.alignment_tab, 'Y_c') and hasattr(self.fitpoints_tab, 'Y_dat_initialguess_calibrated')):
            QMessageBox.critical(
                self.main_window, "Error",
                "This action is missing required previous steps. Red: Missing steps. Yellow: Not done. Green: Done"
            )
            print("Error: Missing required previous steps for calibration conversion")
            return

        if self.select_calibration_tab.G_cal_setting != 2:
            QMessageBox.critical(
                self.main_window, "Error",
                "Calibration data must be in resistivity mode (G_cal_setting=2) to convert to charge carriers"
            )
            print(f"Error: G_cal_setting={self.select_calibration_tab.G_cal_setting}, must be 2 for conversion")
            return

        X_cal, Y_cal = self.alignment_tab.apply_parameters_to_data(
            self.alignment_tab.X_c, self.alignment_tab.Y_c, self.alignment_tab.borders_cal, self.alignment_tab.cal_is_flipped
        )
        X_dat, Y_dat = self.alignment_tab.apply_parameters_to_data(
            self.alignment_tab.X_data, self.alignment_tab.Y_data, self.alignment_tab.borders_data, self.alignment_tab.data_is_flipped
        )
        X_cal, Y_cal, X_dat, Y_dat = self.alignment_tab.apply_lin_offset(
            X_cal, Y_cal, X_dat, Y_dat, self.alignment_tab.best_m, self.alignment_tab.best_t
        )
        self.fitpoints_tab.Y_plateaus_cal = self.fitpoints_tab.Y_plateaus_cal
        Y_plateaus_dat = self.fitpoints_tab.Y_plateaus_dat

        QMessageBox.information(
            self.main_window, "Info",
            f"The mobility model will assume {self.select_calibration_tab.G_dopant_type} doping. "
            f"This can be changed in the calibration import tab"
        )
        self.fitpoints_tab.Y_plateaus_cal_conv = np.log10(self.convert_rho_to_N(np.power(10., self.fitpoints_tab.Y_plateaus_cal)))
        Y_cal_conv = np.log10(self.convert_rho_to_N(np.power(10., Y_cal)))
        print(f"PyQt - calibration_convert: Y_plateaus_cal_conv min={np.min(self.fitpoints_tab.Y_plateaus_cal_conv):.3f}, max={np.max(self.fitpoints_tab.Y_plateaus_cal_conv):.3f}")
        print(f"PyQt - calibration_convert: Y_cal_conv min={np.min(Y_cal_conv):.3f}, max={np.max(Y_cal_conv):.3f}")
        self.select_calibration_tab.G_cal_setting = 1
        self.calibration_convert_metadata[1] = True

        self.figure_calibration_curve.clear()
        ax = self.figure_calibration_curve.add_subplot(111)
        ax.plot(np.power(10., self.fitpoints_dat_opt), np.power(10., self.fitpoints_tab.Y_plateaus_cal_conv),
                color='r', ls='-', label="calibration (optimized)", zorder=1)
        self.draw_xlabel(Quantity="Data", is_log=False)
        self.redraw_calibration_curve_init(ax, X_cal, Y_cal_conv, Y_plateaus_dat, self.fitpoints_tab.Y_plateaus_cal_conv, mode="Calibration")
        ax.legend(loc='best', fontsize=10)
        self.figure_calibration_curve.tight_layout()
        self.canvas_calibration_curve.draw_idle()
        print("PyQt - Calibration curve plotted with converted data")

        self.Y_cal_conv = Y_cal_conv
        self.select_calibration_tab.G_cal_setting = 2
        print(f"PyQt - calibration_convert: G_cal_setting set to {self.select_calibration_tab.G_cal_setting}")

        self.ui.Convert_to_Charge_Carriers_Button.setStyleSheet("background-color: green; color: black")
        self.ui.Convert_to_Charge_Carriers_Button.setEnabled(True)
        try:
            self.ui.Save_To_Database_Button.setStyleSheet("background-color: yellow; color: black")
            self.ui.Save_To_Database_Button.setEnabled(True)
            self.ui.Create_excel_File_Button.setStyleSheet("background-color: yellow; color: black")
            self.ui.Create_excel_File_Button.setEnabled(True)
            self.ui.Save_As_Png_Button.setStyleSheet("background-color: yellow; color: black")
            self.ui.Save_As_Png_Button.setEnabled(True)
            self.ui.Apply_To_Gwyddion_File_Button.setStyleSheet("background-color: yellow; color: black")
            self.ui.Apply_To_Gwyddion_File_Button.setEnabled(True)
            print("PyQt - Export buttons updated")
        except AttributeError as e:
            print(f"PyQt - Warning: Export buttons not found - {e}")

    def calibration_start(self):
        """Optimize calibration curve and update plots."""
        print(f"PyQt - calibration_start: Y_plateaus_cal exists: {hasattr(self.fitpoints_tab, 'Y_plateaus_cal')}")
        print(f"PyQt - calibration_start: Y_plateaus_dat exists: {hasattr(self.fitpoints_tab, 'Y_plateaus_dat')}")
        print(f"PyQt - calibration_start: initialguess exists: {hasattr(self.fitpoints_tab, 'initialguess')}")
        if not (hasattr(self.fitpoints_tab, 'Y_plateaus_cal') and hasattr(self.fitpoints_tab, 'Y_plateaus_dat') and
                hasattr(self.fitpoints_tab, 'initialguess')):
            QMessageBox.critical(
                self.main_window, "Error",
                "This action is missing required previous steps. Red: Missing steps. Yellow: Not done. Green: Done"
            )
            print("Error: Missing required previous steps for calibration")
            return

        X_cal, Y_cal = self.alignment_tab.apply_parameters_to_data(
            self.alignment_tab.X_c, self.alignment_tab.Y_c, self.alignment_tab.borders_cal, self.alignment_tab.cal_is_flipped
        )
        X_dat, Y_dat = self.alignment_tab.apply_parameters_to_data(
            self.alignment_tab.X_data, self.alignment_tab.Y_data, self.alignment_tab.borders_data, self.alignment_tab.data_is_flipped
        )
        X_cal, Y_cal, X_dat, Y_dat = self.alignment_tab.apply_lin_offset(
            X_cal, Y_cal, X_dat, Y_dat, self.alignment_tab.best_m, self.alignment_tab.best_t
        )
        print(f"PyQt - calibration_start: X_cal min={np.min(X_cal):.3f}, max={np.max(X_cal):.3f}, len={len(X_cal)}")
        print(f"PyQt - calibration_start: Y_cal min={np.min(Y_cal):.3f}, max={np.max(Y_cal):.3f}, len={len(Y_cal)}")
        print(f"PyQt - calibration_start: X_dat min={np.min(X_dat):.3f}, max={np.max(X_dat):.3f}, len={len(X_dat)}")
        print(f"PyQt - calibration_start: Y_dat min={np.min(Y_dat):.3f}, max={np.max(Y_dat):.3f}, len={len(Y_dat)}")
        print(f"PyQt - calibration_start: Y_plateaus_cal={self.fitpoints_tab.Y_plateaus_cal}")
        print(f"PyQt - calibration_start: initialguess={self.fitpoints_tab.initialguess}")

        interpolation = self.fitpoints_tab.interpolation
        ref_X_dat = self.alignment_tab.ref(X_dat)
        print(f"PyQt - calibration_start: ref_X_dat min={np.min(ref_X_dat):.3f}, max={np.max(ref_X_dat):.3f}, len={len(ref_X_dat)}")
        test_output = interpolation(X_cal, *self.fitpoints_tab.initialguess)
        print(f"PyQt - calibration_start: test_output min={np.min(test_output):.3f}, max={np.max(test_output):.3f}, len={len(test_output)}")

        try:
            popt, pcov = curve_fit(interpolation, X_cal, Y_cal, p0=self.fitpoints_tab.initialguess)
            print(f"PyQt - calibration_start: popt={popt}")
            self.fitpoints_dat_opt = deepcopy(self.fitpoints_tab.Y_plateaus_dat)
            Y_dat_optimized_calibrated = interpolation(X_dat, *popt)
            print(f"PyQt - calibration_start: Y_dat_optimized_calibrated min={np.min(Y_dat_optimized_calibrated):.3f}, max={np.max(Y_dat_optimized_calibrated):.3f}, len={len(Y_dat_optimized_calibrated)}")

            sign_data = np.sign(np.sum(popt[1:]))
            j = 0
            for i in range(len(popt)):
                if i == 0:
                    self.fitpoints_dat_opt[i] = j + popt[i]
                else:
                    self.fitpoints_dat_opt[i] = j + sign_data * abs(popt[i])
                j = self.fitpoints_dat_opt[i]
            print(f"PyQt - calibration_start: fitpoints_dat_opt={self.fitpoints_dat_opt}")
        except Exception as e:
            QMessageBox.critical(self.main_window, "Error", "Fit did not converge. Please change Fitpoints and try again")
            print(f"Error: Fit did not converge - {e}")
            return

        self.popt = popt

        # Overlay
        self.figure_calibration_overlay.clear()
        ax = self.figure_calibration_overlay.add_subplot(111)
        self.redraw_calibration_overlay_init(ax, X_cal, Y_cal, X_dat, self.fitpoints_tab.Y_dat_initialguess_calibrated)
        ax.plot(
            X_dat, np.power(10., Y_dat_optimized_calibrated), color='r', ls='-',
            label="Measurement Data Calibrated with Optimized Calibration Curve", zorder=1
        )
        ax.legend(loc='best', fontsize=10)

        self.draw_ylabel(
            Quantity="Calibration",
            is_log=not self.main_window.select_calibration_tab.scale_cal_data,
            figure=self.figure_calibration_overlay
        )
        ax.set_xlabel("Depth [µm]")
        self.figure_calibration_overlay.tight_layout()
        self.canvas_calibration_overlay.draw_idle()
        print("PyQt - Calibration overlay plotted with optimized curve")

        # Curve
        self.figure_calibration_curve.clear()
        ax3 = self.figure_calibration_curve.add_subplot(111)
        ax3.plot(np.power(10., self.fitpoints_dat_opt), np.power(10., self.fitpoints_tab.Y_plateaus_cal),
                 color='r', ls='-', label="Calibration (optimized)", zorder=1)
        self.redraw_calibration_curve_init(
            ax3, X_cal, Y_cal, self.fitpoints_tab.Y_plateaus_dat, self.fitpoints_tab.Y_plateaus_cal, mode="Calibration"
        )
        ax3.legend(loc='best', fontsize=10)
        self.draw_ylabel(
            Quantity="Calibration",
            is_log=not self.main_window.select_calibration_tab.scale_cal_data,
            figure=self.figure_calibration_curve
        )
        self.draw_xlabel(Quantity="Data", is_log=False)
        self.figure_calibration_curve.tight_layout()
        self.canvas_calibration_curve.draw_idle()
        print("PyQt - Calibration curve plotted with optimized curve")

        self.fitpoints_dat_opt = self.fitpoints_dat_opt
        self.ui.calibration_startt_pushButton.setStyleSheet("background-color: green; color: black")
        self.ui.calibration_startt_pushButton.setEnabled(True)
        try:
            self.ui.Save_To_Database_Button.setStyleSheet("background-color: yellow; color: black")
            self.ui.Save_To_Database_Button.setEnabled(True)
            self.ui.Create_excel_File_Button.setStyleSheet("background-color: yellow; color: black")
            self.ui.Save_As_Png_Button.setStyleSheet("background-color: yellow; color: black")
            self.ui.Apply_To_Gwyddion_File_Button.setStyleSheet("background-color: yellow; color: black")

            if self.select_calibration_tab.G_cal_setting == 2:
                self.ui.Convert_to_Charge_Carriers_Button.setStyleSheet("background-color: yellow; color: black")
                self.ui.Convert_to_Charge_Carriers_Button.setEnabled(True)
            else:
                self.ui.Convert_to_Charge_Carriers_Button.setStyleSheet("background-color: red; color: black")
                self.ui.Convert_to_Charge_Carriers_Button.setEnabled(False)
            print("PyQt - Export and convert buttons updated")
        except AttributeError as e:
            print(f"PyQt - Warning: Export/convert buttons not found - {e}")



    def export_database(self):
        """Save calibration data to database."""
        if not (hasattr(self.fitpoints_tab, 'Y_plateaus_cal') and hasattr(self.fitpoints_tab, 'Y_plateaus_dat') and
                hasattr(self.fitpoints_tab, 'initialguess') and hasattr(self, 'fitpoints_dat_opt')):
            QMessageBox.critical(
                self.main_window, "Error",
                "Missing required data for database export. Complete previous steps."
            )
            print("Error: Missing required data for database export")
            return

        X_cal, Y_cal = self.alignment_tab.apply_parameters_to_data(
            self.alignment_tab.X_c, self.alignment_tab.Y_c, self.alignment_tab.borders_cal, self.alignment_tab.cal_is_flipped
        )
        X_dat, Y_dat = self.alignment_tab.apply_parameters_to_data(
            self.alignment_tab.X_data, self.alignment_tab.Y_data, self.alignment_tab.borders_data, self.alignment_tab.data_is_flipped
        )
        X_cal, Y_cal, X_dat, Y_dat = self.alignment_tab.apply_lin_offset(
            X_cal, Y_cal, X_dat, Y_dat, self.alignment_tab.best_m, self.alignment_tab.best_t
        )

        if self.select_calibration_tab.G_cal_setting == 1:
            res = None
            cc = self.fitpoints_tab.Y_plateaus_cal
        elif self.select_calibration_tab.G_cal_setting == 2:
            res = self.fitpoints_tab.Y_plateaus_cal
            cc = self.fitpoints_tab.Y_plateaus_cal_conv if hasattr(self.fitpoints_tab, 'Y_plateaus_cal_conv') else None  # This used to falsly call self.Y_cal_conv
        elif self.select_calibration_tab.G_cal_setting == 3:
            res = self.fitpoints_tab.Y_plateaus_cal
            cc = None
        else:
            print(f"Error: G_cal_setting={self.select_calibration_tab.G_cal_setting} is invalid")
            return

        quality = self.alignment_tab.quality
        data_path = getattr(self.import_measurement_tab, 'path_data', 'unknown_sample')
        print(f"Using data_path: {data_path}")
        self.data_path = data_path

        db = calibrationset(data_path=data_path, version="v0.5")
        db.Dopant_type = self.select_calibration_tab.G_dopant_type
        db.carrier_type = self.select_calibration_tab.G_carrier_typee
        db.cal_setting = self.select_calibration_tab.G_cal_setting
        db.ref = self.alignment_tab.ref
        db.fill_set(self.alignment_tab.cal_name, quality, X_cal, Y_cal,
                    self.fitpoints_tab.initialguess, res, cc, self.fitpoints_dat_opt)
        db.save_to_database()
        print("PyQt - Database saved successfully")
        save_measurement_settings_to_json(self.main_window)

        try:
            self.ui.Save_To_Database_Button.setStyleSheet("background-color: green; color: black")
            print("PyQt - export_database_pushButton updated to green")
        except AttributeError as e:
            print(f"PyQt - Warning: export_database_pushButton not found - {e}")
    



    def export_excel(self):
        """Export data to Excel file, mimicking PySimpleGUI -export_excel- event."""
        print("Exporting to Excel")
        self.export_excel_metadata = True
        if self.export_excel_metadata:
            try:
                workbook = load_workbook(filename=self.XLS)
                sheet = workbook['Generic 10-step staircase']
    
                for row in range(len(self.fitpoints_dat_opt)):
                    sheet.cell(row=3 + 2 * row, column=4).value = float(self.fitpoints_dat_opt[row])
    
                    if self.main_window.select_calibration_tab.G_cal_setting == 2:
                        if self.calibration_convert_metadata[1] and hasattr(self.main_window.fitpoints_tab, 'Y_plateaus_cal_conv') and self.main_window.fitpoints_tab.Y_plateaus_cal_conv is not None:
                            sheet.cell(row=3 + 2 * row, column=3).value = float(np.power(10., self.main_window.fitpoints_tab.Y_plateaus_cal_conv[row]))
                            sheet.cell(row=3 + 2 * row, column=2).value = float(np.power(10., self.main_window.fitpoints_tab.Y_plateaus_cal[row]))
                        else:
                            sheet.cell(row=3 + 2 * row, column=2).value = float(np.power(10., self.main_window.fitpoints_tab.Y_plateaus_cal[row]))
                    else:
                        sheet.cell(row=3 + 2 * row, column=3).value = float(np.power(10., self.main_window.fitpoints_tab.Y_plateaus_cal[row]))
    
                data_path = getattr(self.import_measurement_tab, 'path_data', 'unknown_sample')
                print(f"Using data_path: {data_path}")
                self.data_path = data_path
    
                # Create output Excel file path by replacing extension with .xlsx
                path_xl = os.path.splitext(data_path)[0] + ".xlsx"
                workbook.save(filename=path_xl)
                workbook.close()
    
                QMessageBox.information(self.main_window, "Success", f"File saved under {path_xl}")
                self.ui.Create_excel_File_Button.setStyleSheet("background-color: green; color: black;")
                print(f"Excel file saved: {path_xl}")
    
            except Exception as e:
                print(f"Error during Excel export: {e}")
                QMessageBox.critical(self.main_window, "Error", "Failed to export Excel file. Check data and file path.")
        else:
            QMessageBox.critical(
                self.main_window, "Error",
                "This action is missing required previous steps. Red: This step is missing previous steps. "
                "Yellow: This step has not been done. Green: This step has been done"
            )
            print("Export Excel failed: Missing prerequisites")

    def save_as_png(self):
        """Save calibration curve as PNG without rerendering UI plot."""
        print("Saving calibration curve as PNG")
        if hasattr(self, 'fitpoints_dat_opt') and self.fitpoints_dat_opt is not None and \
           hasattr(self.fitpoints_tab, 'Y_plateaus_cal') and self.fitpoints_tab.Y_plateaus_cal is not None and \
           hasattr(self.alignment_tab, 'X_c') and self.alignment_tab.X_c is not None and \
           hasattr(self.alignment_tab, 'Y_c') and self.alignment_tab.Y_c is not None and \
           hasattr(self.fitpoints_tab, 'Y_plateaus_dat') and self.fitpoints_tab.Y_plateaus_dat is not None:
            try:
                X_cal, Y_cal = self.alignment_tab.apply_parameters_to_data(
                    self.alignment_tab.X_c, self.alignment_tab.Y_c, self.alignment_tab.borders_cal, self.alignment_tab.cal_is_flipped
                )
                X_dat, Y_dat = self.alignment_tab.apply_parameters_to_data(
                    self.alignment_tab.X_data, self.alignment_tab.Y_data, self.alignment_tab.borders_data, self.alignment_tab.data_is_flipped
                )
                X_cal, Y_cal, X_dat, Y_dat = self.alignment_tab.apply_lin_offset(
                    X_cal, Y_cal, X_dat, Y_dat, self.alignment_tab.best_m, self.alignment_tab.best_t
                )
    
                fig_png = Figure(figsize=1.5 * self.G_canvas_aspect_ratio / self.G_canvas_dpi, dpi=self.G_canvas_dpi)
                ax = fig_png.add_subplot(111)
                ax.plot(np.power(10., self.fitpoints_dat_opt), np.power(10., self.fitpoints_tab.Y_plateaus_cal),
                        color='r', ls='-', label="calibration (optimized)", zorder=1)
                self.redraw_calibration_curve_init(ax, X_cal, Y_cal, self.fitpoints_tab.Y_plateaus_dat, self.fitpoints_tab.Y_plateaus_cal, mode="Calibration")
                self.draw_xlabel(Quantity="Data", is_log=False, figure=fig_png)
                self.draw_ylabel(Quantity="Calibration", is_log=not self.main_window.select_calibration_tab.scale_cal_data, figure=fig_png)
                ax.legend(loc='best', fontsize=10)
                fig_png.tight_layout()
    
                data_path = getattr(self.import_measurement_tab, 'path_data', 'unknown_sample')
                print(f"Using data_path: {data_path}")
                path_save_png = os.path.splitext(data_path)[0] + " - calibration_curve.png"
                fig_png.savefig(path_save_png, dpi=self.G_canvas_dpi)
                print(f"PNG saved: {path_save_png}")
    
                QMessageBox.information(self.main_window, "Success", f"Saved file as {path_save_png}")
                self.ui.Save_As_Png_Button.setStyleSheet("background-color: green; color: black;")
    
            except Exception as e:
                print(f"Error during PNG export: {e}")
                QMessageBox.critical(self.main_window, "Error", "Failed to save PNG. Check data and file path.")
        else:
            QMessageBox.critical(
                self.main_window, "Error",
                "This action is missing required previous steps. Red: This step is missing previous steps. "
                "Yellow: This step has not been done. Green: This step has been done"
            )
            print("Save PNG failed: Missing prerequisites")

    # -------------------------------------------------------------------------
    # Gwyddion helpers
    # -------------------------------------------------------------------------

    def check_for_gwyddion_file(self, file):
        """Validate Gwyddion file and handle .spm conversion."""
        if file:
            file_ext = os.path.splitext(file)[1].lower()
            if file_ext == ".gwy":
                try:
                    obj = gwyfile.load(file)
                    gwyfile.util.get_datafields(obj)
                    return True
                except Exception as e:
                    QMessageBox.critical(self.main_window, "Error", "Gwyddion file is corrupted. Cannot be opened. Process aborted.")
                    print(f"Error: Gwyddion file corrupted - {e}")
                    return False
            elif file_ext == ".spm":
                try:
                    os.startfile(file)
                    new_file = os.path.splitext(file)[0] + ".gwy"
                    while True:
                        QMessageBox.information(
                            self.main_window, "Info",
                            f"You selected a .spm file. Please save it in Gwyddion as:\n{new_file}\nThen continue."
                        )
                        reply = QMessageBox.question(
                            self.main_window, "Confirm",
                            f"Have you saved the .spm file as {new_file}?",
                            QMessageBox.Yes | QMessageBox.No
                        )
                        if reply == QMessageBox.Yes:
                            try:
                                obj = gwyfile.load(new_file)
                                gwyfile.util.get_datafields(obj)
                                return True
                            except Exception as e:
                                reply = QMessageBox.question(
                                    self.main_window, "Error",
                                    "File cannot be read. Try again?",
                                    QMessageBox.Ok | QMessageBox.Cancel
                                )
                                if reply == QMessageBox.Cancel:
                                    print(f"Error: Failed to read .gwy file - {e}")
                                    return False
                        else:
                            reply = QMessageBox.question(self.main_window, "Confirm", "Abort process?",
                                                         QMessageBox.Yes | QMessageBox.No)
                            if reply == QMessageBox.Yes:
                                print("Gwyddion file selection aborted")
                                return False
                except Exception as e:
                    print(f"Error opening .spm file: {e}")
                    return False
            else:
                QMessageBox.critical(self.main_window, "Error", "File is neither .gwy nor .spm. Process terminated.")
                print("Error: Invalid file extension")
                return False
        return False

    def apply_to_gwyddion_file(self):
        """Apply calibration to Gwyddion file data channels."""
        print("Applying calibration to Gwyddion file")
        if hasattr(self, 'fitpoints_dat_opt') and self.fitpoints_dat_opt is not None and \
           hasattr(self.fitpoints_tab, 'Y_plateaus_cal') and self.fitpoints_tab.Y_plateaus_cal is not None and \
           hasattr(self.fitpoints_tab, 'Y_plateaus_dat') and self.fitpoints_tab.Y_plateaus_dat is not None and \
           hasattr(self.fitpoints_tab, 'initialguess') and self.fitpoints_tab.initialguess is not None:
            try:
                QMessageBox.information(
                    self.main_window, "The following steps are:",
                    "The following steps are:\n 1. Use the following input prompt to open the gwyddion file which contains the data which must be calibrated.\n"
                    " 2. Rename Datachannels, which have duplicate names if neccessary.\n 3. Save in Gwyddion!\n"
                    " 4.Enter the names of the Datachannels in the window which popped up.\n"
                    " 5. Accept last popup to close ANY open Gwyddion instance. \n"
                    " 6. A copy of the gwyddion file will be generated in the same folder containing the calibrated data."
                )

                file, _ = QFileDialog.getOpenFileName(self.main_window, "Select Gwyddion File", "", "Gwyddion Files (*.gwy *.spm)")
                if not self.check_for_gwyddion_file(file):
                    return

                if os.path.splitext(file)[1].lower() == ".spm":
                    file = os.path.splitext(file)[0] + ".gwy"

                obj = gwyfile.load(file)
                df = gwyfile.util.get_datafields(obj)
                datachannels = "\n".join(list(df.keys()))

                dialog = QDialog(self.main_window)
                dialog.setWindowTitle("Select Data Channels")
                dialog.setFixedSize(850, 350)
                layout = QVBoxLayout()

                label = QLabel(
                    "List here all channel names, which should be converted to charge carrier concentration via the calibration curve.\n"
                    "The Gwyddion File will be read after you close this window, so it is possible to change channel names and save the Gwyddion File now."
                )
                label.setFont(QFont("Arial", 8, QFont.Bold))
                label.setAlignment(Qt.AlignCenter)
                label.setWordWrap(True)
                layout.addWidget(label)

                main_layout = QHBoxLayout()
                output_layout = QVBoxLayout()
                input_layout = QVBoxLayout()
                mid_layout = QVBoxLayout()

                output_label = QLabel("Datachannels in File\n(duplicates can't be read!)")
                output_text = QTextEdit()
                output_text.setText(datachannels)
                output_text.setReadOnly(True)
                import_button = QPushButton("Copy channels from file")
                output_layout.addWidget(output_label)
                output_layout.addWidget(output_text)
                output_layout.addWidget(import_button)

                input_label = QLabel("Datachannels to be calibrated")
                input_text = QTextEdit()
                check_button = QPushButton("Check input")
                check_button.setStyleSheet("background-color: yellow; color: black")
                check_button.metadata = False
                ok_button = QPushButton("Finished!")
                ok_button.setStyleSheet("background-color: red; color: black")
                input_layout.addWidget(input_label)
                input_layout.addWidget(input_text)
                input_layout.addWidget(check_button)
                input_layout.addWidget(ok_button)

                copy_button = QPushButton("<<<")
                mid_layout.addStretch()
                mid_layout.addWidget(copy_button)
                mid_layout.addStretch()

                main_layout.addLayout(input_layout)
                main_layout.addLayout(mid_layout)
                main_layout.addLayout(output_layout)
                layout.addLayout(main_layout)
                dialog.setLayout(layout)

                def copy_channels():
                    input_text.setText(output_text.toPlainText())
                    check_button.setStyleSheet("background-color: yellow; color: black")
                    check_button.metadata = False
                    ok_button.setStyleSheet("background-color: red; color: black")

                def import_channels():
                    obj = gwyfile.load(file)
                    df = gwyfile.util.get_datafields(obj)
                    output_text.setText("\n".join(list(df.keys())))
                    check_button.setStyleSheet("background-color: yellow; color: black")
                    check_button.metadata = False
                    ok_button.setStyleSheet("background-color: red; color: black")

                def check_channels():
                    obj = gwyfile.load(file)
                    df = gwyfile.util.get_datafields(obj)
                    channels = input_text.toPlainText().splitlines()
                    valid = all(channel in df for channel in channels if channel.strip())
                    if not valid:
                        QMessageBox.critical(dialog, "Error", "One or more datachannels is invalid! Correct input and save Gwyddion file changes.")
                    else:
                        check_button.setStyleSheet("background-color: green; color: black")
                        check_button.metadata = True
                        ok_button.setStyleSheet("background-color: yellow; color: black")

                copy_button.clicked.connect(copy_channels)
                import_button.clicked.connect(import_channels)
                check_button.clicked.connect(check_channels)
                ok_button.clicked.connect(dialog.accept)
                dialog.rejected.connect(lambda: None)

                if dialog.exec_() == QDialog.Accepted and check_button.metadata:
                    obj_read = gwyfile.load(file)
                    df = gwyfile.util.get_datafields(obj_read)
                    obj_write = GwyContainer()
                    channels = input_text.toPlainText().splitlines()
                    channels = [c for c in channels if c.strip()]

                    print("PyQt Debug - Calibration Parameters:")
                    print(f"  fitpoints_dat_opt: {self.fitpoints_dat_opt}")
                    print(f"  Y_plateaus_cal: {self.fitpoints_tab.Y_plateaus_cal}")
                    print(f"  Y_plateaus_dat: {self.fitpoints_tab.Y_plateaus_dat}")
                    print(f"  initialguess: {self.fitpoints_tab.initialguess}")
                    print(f"  G_cal_setting: {self.select_calibration_tab.G_cal_setting}")
                    print(f"  calibration_convert_metadata[1]: {self.calibration_convert_metadata[1]}")
                    if self.select_calibration_tab.G_cal_setting == 2 and self.calibration_convert_metadata[1]:
                        print(f"  Y_plateaus_cal_conv: {self.fitpoints_tab.Y_plateaus_cal_conv}")

                    for i, channel in enumerate(channels):
                        dfi = df[channel]
                        xres = dfi['xres']
                        yres = dfi['yres']
                        datac = self.fitpoints_tab.linint_(dfi['data'], *self.popt).reshape((yres, xres))

                        if self.select_calibration_tab.G_cal_setting == 2 and self.calibration_convert_metadata[1]:
                            interpolation_cc, linint_cc = self.fitpoints_tab.make_func(self.fitpoints_tab.Y_plateaus_cal_conv)
                            datac = linint_cc(dfi['data'], *self.popt).reshape((yres, xres))
                        elif self.select_calibration_tab.G_cal_setting == 2:
                            if i == 0:
                                QMessageBox.information(
                                    self.main_window, "Info",
                                    "Calibration data is in resistivity, but charge carrier conversion was not done. Output will be in resistivity."
                                )

                        obj_write[f"/{i}/data/title"] = f"{channel}_cal"
                        dfi.data = datac
                        dfi.si_unit_z = GwySIUnit([('unitstr', '')])
                        obj_write[f"/{i}/data"] = dfi

                    calibrated_file = os.path.splitext(file)[0] + "_calibrated.gwy"
                    obj_write.tofile(calibrated_file)
                    os.startfile(calibrated_file)
                    print(f"Calibrated Gwyddion file saved: {calibrated_file}")

                    self.ui.Apply_To_Gwyddion_File_Button.setStyleSheet("background-color: green; color: black")
                    QMessageBox.information(self.main_window, "Success", f"Calibrated file saved as {calibrated_file}")

                else:
                    QMessageBox.critical(self.main_window, "Error", "Selection process aborted or not validated.")
                    print("Gwyddion file calibration aborted")

            except Exception as e:
                print(f"Error during Gwyddion file export: {e}")
                QMessageBox.critical(self.main_window, "Error", "Failed to process Gwyddion file. Check file and data.")
        else:
            QMessageBox.critical(
                self.main_window, "Error",
                "This action is missing required previous steps. Red: This step is missing previous steps. "
                "Yellow: This step has not been done. Green: This step has been done"
            )
            print("Gwyddion file export failed: Missing prerequisites")


